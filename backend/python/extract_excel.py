#!/usr/bin/env python3
"""
Excel extraction for Key Reports financial documents.
Replaces the Node.js xlsx-based extraction with openpyxl.

Supports:
  --type profit_loss    → profit_loss_entries rows
  --type balance_sheet  → balance_sheet_entries rows
  --type general_ledger → general_ledger_entries rows
  --type bank_statement → bank_statement_entries rows

Protocol:
  stdin  : raw Excel/CSV file bytes
  stdout : JSON { rows: [...], detected_years: [...] }
  stderr : debug logs (safe to ignore)
  exit 0 : success
  exit 1 : error (stdout contains { error: "..." })
"""

import sys
import io
import re
import argparse
import calendar
from datetime import datetime

try:
    from openpyxl import load_workbook
except ImportError:
    print('{"error":"openpyxl not installed — run: pip install openpyxl","rows":[],"detected_years":[]}')
    sys.exit(1)

import json as _json

from common import (
    lc, cell_str, parse_amount, parse_date,
    extract_year_from_text, extract_years_from_title_rows, extract_year_cols_from_header,
    is_total_row,
    header_score_pl, header_score_gl, header_score_bs,
    find_best_header_row, find_col,
    ACCOUNT_ALIASES, AMOUNT_ALIASES, DATE_ALIASES, REF_ALIASES,
    NAME_ALIASES, ACCTNUM_ALIASES, DEBIT_ALIASES, CREDIT_ALIASES,
    DESC_ALIASES, CLASS_ALIASES, TYPE_ALIASES, BALANCE_ALIASES, SPLIT_ALIASES,
    emit, emit_error,
)

AS_OF_PATTERN = re.compile(
    r'(?:as\s+of\s+)?'
    r'(\w+\s+\d{1,2},?\s*\d{4}'       # "December 31, 2024"
    r'|\d{1,2}[/\-]\d{1,2}[/\-]\d{4}' # "12/31/2024"
    r'|\d{4}-\d{2}-\d{2})',             # "2024-12-31"
    re.IGNORECASE,
)


# ── Workbook helpers ──────────────────────────────────────────────────────────

def load_wb():
    raw = sys.stdin.buffer.read()
    if not raw:
        emit_error('No file data received on stdin')
    try:
        return load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        emit_error(f'Cannot open Excel file: {e}')


def get_rows(wb, pattern=None):
    """
    Return array-of-arrays from the best-matching sheet (or first sheet).
    Each row is a list of raw cell values (None for empty cells).
    """
    target = None
    if pattern:
        for name in wb.sheetnames:
            if re.search(pattern, name, re.IGNORECASE):
                target = name
                break
    ws = wb[target] if target else wb.worksheets[0]
    return [[cell.value for cell in row] for row in ws.iter_rows()]


def get_rows_with_indent(wb, pattern=None):
    """
    Same as get_rows, but also captures each cell's Excel-native alignment
    indent level (cell.alignment.indent — an integer nesting level, e.g. 0
    for "Income", 1 for "Services" nested under it, 2 for "Refunds" nested
    under that).

    CONFIRMED ROOT CAUSE (this is the fix for it): QuickBooks' own "Profit and
    Loss"/"Balance Sheet" Excel exports commonly encode hierarchy via THIS
    cell formatting property, with NO leading whitespace in the cell's text
    value at all (e.g. the cell literally contains "Services", not "
    Services"). The existing indent detection (_cell_indent) only ever
    looked for leading whitespace in the text — for files using this real,
    common export convention, EVERY row's text-based indent was 0, so the
    ancestor stack never retained any parent, parent_path came out empty for
    every account, and downstream P&L section/type resolution had nothing to
    work with. Confirmed live against a real production file.

    Returns (rows, align_indents) — align_indents[i][j] is the alignment
    indent of rows[i][j] (int, defaulting to 0 when unavailable).
    """
    target = None
    if pattern:
        for name in wb.sheetnames:
            if re.search(pattern, name, re.IGNORECASE):
                target = name
                break
    ws = wb[target] if target else wb.worksheets[0]
    rows, align_indents = [], []
    for row in ws.iter_rows():
        rows.append([cell.value for cell in row])
        row_indents = []
        for cell in row:
            indent = 0
            try:
                if cell.alignment and cell.alignment.indent:
                    indent = int(cell.alignment.indent)
            except Exception:
                indent = 0
            row_indents.append(indent)
        align_indents.append(row_indents)
    return rows, align_indents


def dbg(msg):
    print(f'[extract_excel] {msg}', file=sys.stderr, flush=True)


# ── PROFIT & LOSS ─────────────────────────────────────────────────────────────

# Same literal-header-text vocabulary as profitLossExtractionService.js's
# SECTION_HEADER_PATTERNS/matchSectionHeader — recognizes the document's OWN
# bare section-header labels, never an account-name keyword rule.
_PL_SECTION_HEADER_PATTERNS = [
    ('revenue', re.compile(r'^(income|revenue|sales|total income|total revenue)$', re.IGNORECASE)),
    ('cost_of_sales', re.compile(r'^(cost of goods sold|cost of sales|cogs|total cost of goods sold|total cost of sales)$', re.IGNORECASE)),
    ('operating_expenses', re.compile(r'^(expenses|expense|operating expenses|total expenses|total operating expenses)$', re.IGNORECASE)),
    ('other_income', re.compile(r'^(other income|other revenue|interest income|interest earned|financial income|extraordinary income|total other income|total other revenue|net other income)$', re.IGNORECASE)),
    ('other_expense', re.compile(r'^(other expense|other expenses|financial expense|financial expenses|extraordinary expense|extraordinary expenses|total other expense|total other expenses)$', re.IGNORECASE)),
]


def match_section_header(label):
    # CONFIRMED BUG (fixed here): this used to `return True` instead of the
    # actual key string ('revenue'/'cost_of_sales'/...), and no P&L data row
    # ever got a `.section` value at all — every account extracted via this
    # Python path reached chartOfAccountsService.js with section=None,
    # producing account_type=NULL for every single P&L account regardless of
    # header nesting. Mirrors profitLossExtractionService.js's
    # matchSectionHeader exactly (same patterns, same key names).
    norm = str(label or '').strip()
    if not norm:
        return None
    for key, pattern in _PL_SECTION_HEADER_PATTERNS:
        if pattern.match(norm):
            return key
    return None


# CONFIRMED ROOT CAUSE (see match_section_header's comment): every row already
# carries its own real ancestor chain (`parent_path`, built from the
# document's own indentation). Walking that chain from the ROOT (outermost)
# downward and taking the FIRST label that matches one of the fixed anchors
# is self-contained per row — correct regardless of how many unrecognized,
# company-specific intermediate headers (e.g. "Payroll Expenses", "Store
# Expenses") sit between the leaf and its true section, since non-matching
# labels are simply skipped over, never assigned a type of their own. No new
# patterns, no hardcoded company-specific names — mirrors
# profitLossExtractionService.js's sectionFromAncestry exactly.
def section_from_ancestry(parent_path):
    for label in parent_path or []:
        key = match_section_header(label)
        if key:
            return key
    return None


def extract_profit_loss(wb):
    rows, align_indents = get_rows_with_indent(wb, r'income|profit|p&l|pl|earnings|revenue')
    if len(rows) < 2:
        emit_error('No data rows found in P&L sheet')

    header_idx, score = find_best_header_row(rows, header_score_pl)
    dbg(f'Header row: {header_idx} (score={score}): {[cell_str(c) for c in rows[header_idx]][:8]}')

    header_row = rows[header_idx]

    # Extract years from title area — INCLUSIVE of header_idx (fixes JS year-scan bug)
    title_years = extract_years_from_title_rows(rows, header_idx)
    year_cols   = extract_year_cols_from_header(header_row)

    # Determine which year columns to emit
    if year_cols:
        fiscal_years = year_cols                            # multi-year columns
    else:
        guess = title_years[-1] if title_years else datetime.now().year
        fiscal_years = [{'col_idx': -1, 'year': guess}]   # single-year

    acct_idx = max(0, find_col(header_row, ACCOUNT_ALIASES))
    dbg(f'Account col={acct_idx}, year cols={fiscal_years}')

    # Which hierarchy signal this SHEET actually uses — checked once, not
    # per-row: a real document encodes nesting ONE way, not a mix of both.
    # See get_rows_with_indent's doc comment for why alignment indent must be
    # checked at all (QuickBooks-style exports carry no leading whitespace).
    uses_alignment_indent = any(
        acct_idx < len(align_indents[r]) and align_indents[r][acct_idx] > 0
        for r in range(header_idx + 1, len(rows))
    )
    dbg(f'Hierarchy signal: {"cell alignment indent" if uses_alignment_indent else "leading whitespace in text"}')

    result = []
    # Real multi-level hierarchy read from the document's own indentation —
    # same stack discipline used in extract_balance_sheet.
    ancestor_stack = []  # list of (indent, label)

    # No period-caption guard here, unlike extract_balance_sheet: this function
    # reads its period columns straight off header_row itself
    # (extract_year_cols_from_header above), so the captions ARE the header row
    # and the loop below — which starts at header_idx + 1 — can never reach
    # them. The balance-sheet extractor scans for its caption row separately and
    # does need the guard; see period_header_idx there.
    for row_idx in range(header_idx + 1, len(rows)):
        row = rows[row_idx]
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        raw_cell = row[acct_idx] if acct_idx < len(row) else None
        raw_name = cell_str(raw_cell)
        if not raw_name:
            continue

        if uses_alignment_indent:
            indent = align_indents[row_idx][acct_idx] if acct_idx < len(align_indents[row_idx]) else 0
        else:
            indent = _cell_indent(raw_cell)
        while ancestor_stack and ancestor_stack[-1][0] >= indent:
            ancestor_stack.pop()
        parent_path = [label for _, label in ancestor_stack]

        account_name = raw_name.lstrip()   # strip leading indent spaces
        is_total     = is_total_row(account_name)
        is_subtotal  = bool(re.match(
            r'^(gross profit|net operating income|net other income|operating income|net income|net loss)$',
            account_name, re.IGNORECASE,
        ))

        # A bare header row (no amount anywhere, not a total line) still
        # becomes an ancestor for whatever is nested under it — preserved
        # (never silently dropped) so the full document hierarchy reaches the
        # COA generator, same node_type convention as extract_balance_sheet.
        # P&L has no persisted table, so this is purely additive.
        has_any_amount = any(
            parse_amount(row[yc['col_idx']] if yc['col_idx'] < len(row) else None) is not None
            for yc in year_cols
        ) if year_cols else any(
            parse_amount(v) is not None for v in row[acct_idx + 1:]
        )
        if not has_any_amount and not is_total:
            header_key = match_section_header(account_name)
            header_years = [yc['year'] for yc in year_cols] if year_cols else [fiscal_years[0]['year']]
            for year in header_years:
                result.append({
                    'account_name': account_name,
                    'account_type': None,
                    'section':      header_key or section_from_ancestry(parent_path),
                    'parent_path':  parent_path,
                    'amount':       0,
                    'fiscal_year':  year,
                    'is_total':     False,
                    'is_header':    True,
                    'node_type':    'hierarchy_section' if header_key else 'hierarchy_group',
                })
            ancestor_stack.append((indent, account_name))
            continue

        node_type = 'subtotal' if is_subtotal else ('total' if is_total else 'account')
        section = section_from_ancestry(parent_path)

        if year_cols:
            for yc in year_cols:
                c = yc['col_idx']
                amt = parse_amount(row[c] if c < len(row) else None)
                if amt is None and not is_total:
                    continue
                result.append({
                    'account_name': account_name,
                    'account_type': None,
                    'section':      section,
                    'parent_path':  parent_path,
                    'amount':       amt if amt is not None else 0,
                    'fiscal_year':  yc['year'],
                    'is_total':     is_total,
                    'is_header':    False,
                    'node_type':    node_type,
                })
        else:
            year = fiscal_years[0]['year']
            # Walk right-to-left to find the last numeric value
            amount = None
            for c in range(len(row) - 1, acct_idx, -1):
                v = parse_amount(row[c])
                if v is not None:
                    amount = v
                    break
            if amount is None:
                continue
            result.append({
                'account_name': account_name,
                'account_type': None,
                'section':      section,
                'parent_path':  parent_path,
                'amount':       amount,
                'fiscal_year':  year,
                'is_total':     is_total,
                'is_header':    False,
                'node_type':    node_type,
            })
        # Leaf/total/subtotal rows are never pushed onto the ancestor stack —
        # see extract_balance_sheet's identical fix for the full explanation.
        # Only the header/group branch above (continue'd before reaching here)
        # is allowed to remain a parent for later rows.

    detected_years = sorted({r['fiscal_year'] for r in result})
    dbg(f'Extracted {len(result)} rows, years={detected_years}')
    return {'rows': result, 'detected_years': detected_years}


# ── BALANCE SHEET ─────────────────────────────────────────────────────────────

_PERIOD_MONTHS = {
    'jan': 1, 'january': 1, 'feb': 2, 'february': 2, 'mar': 3, 'march': 3,
    'apr': 4, 'april': 4, 'may': 5, 'jun': 6, 'june': 6, 'jul': 7, 'july': 7,
    'aug': 8, 'august': 8, 'sep': 9, 'sept': 9, 'september': 9, 'oct': 10,
    'october': 10, 'nov': 11, 'november': 11, 'dec': 12, 'december': 12,
}


def parse_period_header(text):
    """A column header naming ONE reporting period -> (year, month), else None.

    Recognizes the forms a monthly statement column actually uses -- "Jan 2024",
    "January 2024", "Jan-24", "Jan 31, 2024". A "Total" column, a blank, or any
    non-period label returns None so it is never mistaken for a period.
    """
    t = str(text or '').strip().lower()
    if not t or 'total' in t:
        return None
    name = re.search(r'([a-z]{3,9})', t)
    month = _PERIOD_MONTHS.get(name.group(1)) if name else None
    if not month:
        return None
    # A 4-digit year wins outright. Matching the first number instead would read
    # the DAY as the year in the "Jan 31, 2024" form (-> 2031).
    y4 = re.search(r'(\d{4})', t)
    if y4:
        year = int(y4.group(1))
    else:
        y2 = re.search(r'[\s\-/.,](\d{2})(?!\d)', t)
        if not y2:
            return None
        year = 2000 + int(y2.group(1))
    if year < 1900 or year > 2200:
        return None
    return (year, month)


def month_end_date(year, month):
    return f'{year:04d}-{month:02d}-{calendar.monthrange(year, month)[1]:02d}'


def parse_as_of_date(text):
    m = AS_OF_PATTERN.search(str(text or ''))
    if not m:
        return None
    s = m.group(1).strip().rstrip(',')
    for fmt in ('%B %d %Y', '%b %d %Y', '%B %d, %Y', '%b %d, %Y',
                '%m/%d/%Y', '%Y-%m-%d', '%d/%m/%Y', '%m-%d-%Y'):
        try:
            return datetime.strptime(s, fmt).strftime('%Y-%m-%d')
        except ValueError:
            pass
    return None


def _cell_indent(v):
    """Count leading whitespace characters on a RAW (unstripped) cell value —
    tabs count as 4 — before cell_str() strips it. This is the only place the
    document's own indentation (how QuickBooks/Xero/Sage exports encode
    "Checking is nested under Bank Accounts") survives; mirrors the identical
    heuristic in the JS extractors (balanceSheetExtractionService.js /
    profitLossExtractionService.js)."""
    s = str(v) if v is not None else ''
    m = re.match(r'^[ \t]*', s)
    return len(m.group(0).replace('\t', '    ')) if m else 0


# Balance Sheet section-header vocabulary — mirrors
# balanceSheetExtractionService.js's inferSection exactly (whole-word-set
# match, never a substring test on the whole string). common.infer_section is
# intentionally NOT reused for this: it is a plain substring matcher shared
# with the PDF text/OCR fallback paths (which have no indentation/parent_path
# signal to derive a section from and are out of scope here), and substring
# matching on this closed vocabulary would misfire the same way JS's old
# inferSection did on "Capital One Credit Card" (matched "capital").
_BS_HEADER_WORDS = {
    'total', 'current', 'fixed', 'long', 'term', 'other', 'and',
    'asset', 'assets', 'liability', 'liabilities',
}
_BS_EQUITY_WORDS = {
    'equity', 'capital', 'owner', 'owners', 'member', 'members',
    'stockholder', 'stockholders', 'partner', 'partners',
}
_BS_HEADER_WORDS |= _BS_EQUITY_WORDS


def infer_bs_header_section(account_name):
    n = lc(account_name).replace("'", '').strip()
    if not n:
        return None
    words = [w for w in re.sub(r'[^a-z\s]', ' ', n).split() if w]
    if not words or not all(w in _BS_HEADER_WORDS for w in words):
        return None

    has_asset = any(w in ('asset', 'assets') for w in words)
    has_liability = any(w in ('liability', 'liabilities') for w in words)
    has_equity = any(w in _BS_EQUITY_WORDS for w in words)

    if has_asset:
        return 'assets'
    if has_liability and has_equity:
        # Combined "Liabilities and Equity" style header — resolve by which
        # marker word appears first (mirrors JS inferSection's tie-break).
        liab_idx = n.find('liabilit')
        equity_positions = [n.find(w) for w in _BS_EQUITY_WORDS if n.find(w) != -1]
        equity_idx = min(equity_positions) if equity_positions else -1
        if equity_idx == -1:
            return 'liabilities'
        return 'equity' if equity_idx < liab_idx else 'liabilities'
    if has_liability:
        return 'liabilities'
    if has_equity:
        return 'equity'
    return None


# CONFIRMED ROOT CAUSE: `section` was derived from `current_section`, a flat
# variable set only when a recognized header line was seen, and never
# rescoped when the ancestor stack popped back past that header — an Equity
# account nested directly under a combined "Liabilities and Equity" umbrella
# (with no further explicit "Equity" sub-header between it and the umbrella)
# silently kept the stale "Liabilities" section left over from an earlier
# sibling branch (e.g. "Liabilities" > "Accounts Payable"). Every row already
# carries its own real ancestor chain (`parent_path`, built from the
# document's own indentation) — walking that chain from the NEAREST ancestor
# toward the root (the REVERSE of section_from_ancestry's root-first order
# for P&L, which has no umbrella-then-branch ambiguity to resolve) and taking
# the first label that resolves via infer_bs_header_section is self-contained
# per row, and correctly prefers a specific "Equity"/"Liabilities" sub-header
# over the ambiguous umbrella "Liabilities and Equity" heading that may sit
# above it in the same path.
def bs_section_from_ancestry(parent_path):
    for label in reversed(parent_path or []):
        key = infer_bs_header_section(label)
        if key:
            return key
    return None


def extract_balance_sheet(wb):
    rows, align_indents = get_rows_with_indent(wb, r'balance\s*sheet|bs\b')
    if len(rows) < 2:
        emit_error('No data rows found in Balance Sheet')

    # Scan first 20 rows for as_of_date and fiscal year
    as_of_date       = None
    title_fiscal_year = None
    for i in range(min(20, len(rows))):
        row_text = ' '.join(cell_str(c) for c in rows[i])
        if not as_of_date:
            d = parse_as_of_date(row_text)
            if d:
                as_of_date        = d
                title_fiscal_year = int(d[:4])
        if not title_fiscal_year:
            y = extract_year_from_text(row_text)
            if y:
                title_fiscal_year = y

    # Score-based header detection — FIXED vs JS 'last matching row' bug
    header_idx, score = find_best_header_row(rows, header_score_bs, max_scan=20)
    dbg(f'BS header row: {header_idx} (score={score}), as_of={as_of_date}')

    if not as_of_date:
        year       = title_fiscal_year or datetime.now().year
        as_of_date = f'{year}-12-31'
    fiscal_year = title_fiscal_year or int(as_of_date[:4])

    header_row = rows[header_idx]
    acct_idx   = max(0, find_col(header_row, ACCOUNT_ALIASES))

    # CONFIRMED ROOT CAUSE (fixed here): a Balance Sheet exported with ONE COLUMN
    # PER MONTH ("Jan 2024" ... "Dec 2024") was collapsed to a single period --
    # the amount scan below walks right-to-left and stops at the first value, so
    # only the LAST column survived, under the sheet's single as_of_date.
    #
    # Confirmed live on a 12-column export: balance_sheet_entries held exactly
    # one as_of_date per year ("2024-12-30", "2025-12-30", "2026-07-26"). With no
    # monthly document data to read, generateMonthlyBs fell through to
    # GL-derived snapshots for every month, and those do not reconcile to the
    # document -- January 2024 rendered Total Assets 283,931.50 against the
    # uploaded 293,161.70, and Accrued Revenue came out NEGATIVE (-83,830.90
    # against 44,279.84).
    #
    # Each period column is now emitted as its own row, dated to that month's
    # last day. A single-period sheet finds fewer than two period columns and
    # keeps the original behaviour exactly.
    # Scanned independently of header_idx rather than off header_row: the
    # score-based header detector legitimately locks onto the title line ("As of
    # Dec 31, 2024" scored highest on a real file, putting header_idx two rows
    # ABOVE the month captions), and re-tuning it would move the data start row
    # for every statement. The period captions are whichever row yields the most
    # parseable periods.
    period_cols = []
    period_header_idx = None
    for i in range(min(20, len(rows))):
        cands = []
        for c in range(acct_idx + 1, len(rows[i])):
            parsed = parse_period_header(cell_str(rows[i][c]))
            if parsed:
                cands.append((c, parsed[0], parsed[1]))
        if len(cands) > len(period_cols):
            period_cols = cands
            period_header_idx = i
    is_multi_period = len(period_cols) >= 2
    # Structural rows (section headers / grouping labels) carry no figures, so
    # they are emitted ONCE rather than per period. In multi-period mode they
    # must still be dated to the LAST period, never to the sheet's own as_of
    # line: that line can fall on a different day (a real file reads "As of Jul
    # 27, 2026" while its final column is July), which would add a phantom
    # as_of_date holding nothing but zero-amount structural rows. generateYearlyBs
    # picks the LATEST as_of_date for the year, so it would have selected that
    # phantom date and rendered the whole year as zeros.
    structural_as_of = (
        month_end_date(period_cols[-1][1], period_cols[-1][2]) if is_multi_period else as_of_date
    )
    # CONFIRMED ROOT CAUSE (fixed here) of "the December <year> Balance Sheet is
    # missing": in multi-period mode the structural rows were stamped with the
    # SHEET's title-derived `fiscal_year` while the account rows were correctly
    # stamped with each period's own year (`py`). A monthly export whose columns
    # cross a calendar year, or whose title line names a different year from its
    # columns, therefore wrote structural rows into a fiscal year that has no
    # account rows at all -- and generateYearlyBs (financialStatementService)
    # selects the LATEST as_of_date for a fiscal year, so it could land on that
    # structural-only date and render the whole year as zeros.
    #
    # The structural rows belong to the same period `structural_as_of` names, so
    # the fiscal year must be read from that date, never from the title.
    structural_fiscal_year = (
        period_cols[-1][1] if is_multi_period else fiscal_year
    )
    if is_multi_period:
        dbg(f'BS multi-period: {len(period_cols)} period column(s) '
            f'{period_cols[0][1]}-{period_cols[0][2]:02d} .. {period_cols[-1][1]}-{period_cols[-1][2]:02d}')

    # Which hierarchy signal this SHEET actually uses — checked once, not per
    # row (see extract_profit_loss's identical check and get_rows_with_indent's
    # doc comment): a QuickBooks-style export encodes nesting via the cell's
    # own alignment.indent property, with ZERO leading whitespace in the text
    # itself — _cell_indent alone would then read 0 for every row, flattening
    # the entire document hierarchy (confirmed live: this exact bug left every
    # Balance Sheet account's parent_path empty, collapsing Level 3/4 down to
    # the leaf account name instead of "Current Assets"/"Fixed Assets" etc.).
    uses_alignment_indent = any(
        acct_idx < len(align_indents[r]) and align_indents[r][acct_idx] > 0
        for r in range(header_idx + 1, len(rows))
    )
    dbg(f'BS hierarchy signal: {"cell alignment indent" if uses_alignment_indent else "leading whitespace in text"}')

    current_section = None
    result = []
    # Real multi-level hierarchy read from the document's own indentation —
    # same stack discipline as the JS extractors: a row whose indent is <= the
    # stack's top pops it, leaving exactly that row's real ancestor chain.
    ancestor_stack = []  # list of (indent, label)

    for row_idx in range(header_idx + 1, len(rows)):
        # The month captions ("Jan 2024", "Feb 2024", ...) are not an account
        # row. They can legitimately fall BELOW header_idx — the score-based
        # detector locks onto the title line on real files, putting header_idx
        # above the captions (see period_header_idx's comment above) — so this
        # loop can and does reach them. Most exports leave the caption row's
        # account column blank, which the empty-name check below already drops,
        # but one that labels it (e.g. "Account") would otherwise be emitted as
        # a phantom account carrying every period's header text.
        if row_idx == period_header_idx:
            continue
        row = rows[row_idx]
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        raw_cell = row[acct_idx] if acct_idx < len(row) else None
        raw_name = cell_str(raw_cell)
        if not raw_name:
            continue

        if uses_alignment_indent:
            indent = align_indents[row_idx][acct_idx] if acct_idx < len(align_indents[row_idx]) else 0
        else:
            indent = _cell_indent(raw_cell)
        while ancestor_stack and ancestor_stack[-1][0] >= indent:
            ancestor_stack.pop()
        parent_path = [label for _, label in ancestor_stack]

        section_match = infer_bs_header_section(raw_name)

        # Walk right-to-left for amount
        amount = None
        period_amounts = []
        if is_multi_period:
            for c, py, pm in period_cols:
                v = parse_amount(row[c]) if c < len(row) else None
                if v is not None:
                    period_amounts.append((py, pm, v))
            # `amount` still drives the structural branches below: a row with no
            # value in ANY period is a header/group row, exactly as before.
            if period_amounts:
                amount = period_amounts[-1][2]
        else:
            for c in range(len(row) - 1, acct_idx, -1):
                v = parse_amount(row[c])
                if v is not None:
                    amount = v
                    break

        # Pure, RECOGNIZED section header row (no amount) — never a postable
        # account: filterRowsBeforeInsertion (extractionService.base.js)
        # strips it via hierarchy_level=0 before it reaches balance_sheet_entries.
        if section_match and amount is None:
            current_section = raw_name
            result.append({
                'account_name':    raw_name,
                'account_number':  None,
                'section':         section_match,
                'parent_path':     parent_path,
                'amount':          0,
                'as_of_date':      structural_as_of,
                'fiscal_year':     structural_fiscal_year,
                'is_total':        False,
                'is_section_header': True,
                'node_type':       'hierarchy_section',
            })
            ancestor_stack.append((indent, raw_name))
            continue

        if amount is None:
            # An UNRECOGNIZED intermediate grouping label (e.g. "Bank
            # Accounts") — no section keyword match, no amount. Not itself a
            # postable account, but a real ancestor for whatever is nested
            # more deeply under it (the whole point of reading indentation
            # instead of only fixed keywords) — still emitted so the full
            # document hierarchy is never silently discarded, but
            # filterRowsBeforeInsertion strips it the same way as a
            # recognized section header (same hierarchy_level=0 signal).
            result.append({
                'account_name':    raw_name,
                'account_number':  None,
                'section':         bs_section_from_ancestry(parent_path) or (infer_bs_header_section(current_section) if current_section else None),
                'parent_path':     parent_path,
                'amount':          0,
                'as_of_date':      structural_as_of,
                'fiscal_year':     structural_fiscal_year,
                'is_total':        False,
                'is_section_header': False,
                'node_type':       'hierarchy_group',
            })
            ancestor_stack.append((indent, raw_name))
            continue

        account_name = raw_name.lstrip()
        is_total = is_total_row(account_name)
        row_section = bs_section_from_ancestry(parent_path) or (infer_bs_header_section(current_section) if current_section else infer_bs_header_section(account_name))
        # One row per reporting period the document actually states a figure
        # for. A blank cell is simply not a data point for that month and is
        # skipped rather than written as a zero.
        emit_periods = (
            [(py, pm, v) for py, pm, v in period_amounts]
            if is_multi_period
            else [(fiscal_year, None, amount)]
        )
        for py, pm, v in emit_periods:
            result.append({
                'account_name':     account_name,
                'account_number':   None,
                'section':          row_section,
                'parent_path':      parent_path,
                'amount':           v,
                'as_of_date':       month_end_date(py, pm) if pm else as_of_date,
                'fiscal_year':      py,
                'is_total':         is_total,
                'is_section_header': False,
                'node_type':        'total' if is_total else 'account',
            })
        # CONFIRMED ROOT CAUSE (fixed here): a leaf/total row — one that carries
        # its own posted amount — must NEVER be pushed onto the ancestor stack.
        # Only a structural header/group row (no amount anywhere on the line,
        # handled in the two branches above) can be a real parent for what
        # follows. Previously every row was pushed unconditionally, so whenever
        # two sibling accounts' indent readings weren't perfectly monotonic
        # (common indent noise in real Excel exports), the first sibling was
        # silently retained as the "parent" of the second — confirmed live:
        # "Mitch Greene Distribution" nested under an unrelated "Eugene G &
        # Arnold G" equity account, and "22110 Garnishment Payable" nested
        # under sibling "22100 Employee Expense Payable" (11 of 58 real rows
        # in one document). Level 3+ is built entirely from parent_path, so
        # this was the direct cause of intermittent, algorithmic (not
        # client-specific) Level 3/4 corruption. Do not re-add this push.

    # CONFIRMED ROOT CAUSE (fixed here) of the client's "the December <year>
    # Balance Sheet is missing" report.
    #
    # This was `[fiscal_year]` — the SHEET's single title-derived year — while
    # every account row above is correctly stamped with its OWN period's year
    # (`py`, see the emit_periods loop). A multi-period Balance Sheet whose
    # columns span more than one calendar year (a monthly export running e.g.
    # Dec of one year through Dec of the next, or one whose title line names a
    # different year from its columns) therefore reported exactly ONE detected
    # year even though rows for both years were inserted.
    #
    # detected_years is not cosmetic: extractionService.base.js persists it to
    # key_report_document_mappings.metadata.detectedYears, and
    # keyReportValidationService.resolveMappingYears/collectYears reads THAT to
    # decide which fiscal years a version has a Balance Sheet for. The year whose
    # rows existed but whose number never made it into this list was reported to
    # the user as having no Balance Sheet at all.
    #
    # Derived from the rows actually emitted, which is what
    # profitLossExtractionService.js already does on the P&L side
    # (`[...new Set(rows.map(r => r.fiscal_year))]`) — the two extractors now
    # agree instead of one reporting per-period years and the other a title year.
    detected_years = sorted({
        r['fiscal_year'] for r in result
        if isinstance(r.get('fiscal_year'), int) and 1900 <= r['fiscal_year'] <= 2100
    })
    dbg(f'Extracted {len(result)} BS rows, years={detected_years}')
    return {'rows': result, 'detected_years': detected_years}


# ── GENERAL LEDGER ────────────────────────────────────────────────────────────

def extract_general_ledger(wb):
    """
    Extract every row from the GL sheet, preserving source structure.

    Row types emitted:
      ACCOUNT_HEADER    — section header row (e.g. "Business Checking (7454)")
      BEGINNING_BALANCE — "Beginning Balance" sentinel row
      TRANSACTION       — dated transaction row
      TOTAL_ROW         — "Total …" summary row
    """
    rows = get_rows(wb, r'general\s*ledger|gl\b|transaction')
    if len(rows) < 2:
        emit_error('No data rows found in GL file')

    header_idx, score = find_best_header_row(rows, header_score_gl, max_scan=25)
    if score < 2:
        dbg('No strong GL header found — using row 0')
        header_idx = 0

    header_row = rows[header_idx]
    dbg(f'GL header row: {header_idx} (score={score}): {[cell_str(c) for c in header_row][:10]}')

    col = {
        'date':        find_col(header_row, DATE_ALIASES),
        'ref':         find_col(header_row, REF_ALIASES),
        'name':        find_col(header_row, NAME_ALIASES),
        'account':     find_col(header_row, ACCOUNT_ALIASES),
        'account_num': find_col(header_row, ACCTNUM_ALIASES),
        'debit':       find_col(header_row, DEBIT_ALIASES),
        'credit':      find_col(header_row, CREDIT_ALIASES),
        'amount':      find_col(header_row, AMOUNT_ALIASES),
        'desc':        find_col(header_row, DESC_ALIASES),
        'class_':      find_col(header_row, CLASS_ALIASES),
        'type':        find_col(header_row, TYPE_ALIASES),
        'balance':     find_col(header_row, BALANCE_ALIASES),
        'split':       find_col(header_row, SPLIT_ALIASES),
    }
    dbg(f'GL col map: {col}')

    # Detect QBO "by Account" format: no Account column in header; the account
    # name appears as a standalone section-header row above its transactions.
    is_by_account = col['account'] < 0 and col['account_num'] < 0

    # If date column not found, probe first data rows
    if col['date'] < 0:
        for c in range(min(5, len(header_row))):
            for r_idx in range(header_idx + 1, min(header_idx + 6, len(rows))):
                if r_idx < len(rows) and c < len(rows[r_idx]):
                    if parse_date(rows[r_idx][c]):
                        col['date'] = c
                        break
            if col['date'] >= 0:
                break
    if col['date'] < 0:
        emit_error('Could not detect date column in GL file')

    def get(row, key):
        c = col[key]
        return row[c] if c >= 0 and c < len(row) else None

    def last_numeric(row):
        """Return the last parseable numeric value in row, or None."""
        for v in reversed(row):
            a = parse_amount(v)
            if a is not None:
                return a
        return None

    current_account_section = None
    current_fiscal_year     = None
    result = []

    # header_idx is 0-based; Excel rows are 1-based. Row header_idx+2 is the
    # first data row (1 for header, 1 for 0→1 conversion).
    for offset, row in enumerate(rows[header_idx + 1:]):
        excel_row_num = header_idx + 2 + offset   # 1-based source row number

        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        raw_row_json = _json.dumps([cell_str(v) for v in row], default=str)
        first_cell   = cell_str(row[0] if row else None)
        first_lc     = first_cell.lower()
        date_str     = parse_date(get(row, 'date'))

        # CONFIRMED ROOT CAUSE (fixed here): the "Beginning Balance" sentinel was
        # looked up in column 0 only. A very common export layout indents the
        # ledger grid by one column -- the ACCOUNT HEADING sits alone in column 0
        # and every detail row (including the "Beginning Balance" label) starts
        # at the detected account column. On such a file row[0] is empty for that
        # row, so the branch below never matched: it fell through to the
        # TRANSACTION branch, had no date, and was silently discarded.
        #
        # Measured on a real 3-file export: 52 "Beginning Balance" rows in the
        # workbooks, 0 reaching the database. That destroyed every opening
        # balance AND removed the equity accounts entirely -- they carry an
        # opening balance and no transactions, so with the row gone they produced
        # no GL rows at all and vanished from the Trial Balance. That is exactly
        # what the NO_EQUITY_ACCOUNTS warning was reporting.
        #
        # Deliberately scoped to THIS lookup. `first_cell` must stay column-0
        # only: the ACCOUNT_HEADER branch relies on an outdented heading, and the
        # "Total for X" rows really are in column 0 (verified 82/82 on that
        # export). Widening either would turn ordinary detail rows into bogus
        # section headers or totals.
        label_cell = first_cell or cell_str(get(row, 'account'))
        label_lc   = label_cell.lower()

        # ── BEGINNING BALANCE ─────────────────────────────────────────────────
        if 'beginning balance' in label_lc:
            # An opening balance precedes its section's transactions, so
            # current_fiscal_year is still the PREVIOUS account's year -- and is
            # None outright for the first account in the file. Look ahead to the
            # next dated row instead, which is the year this balance opens.
            bb_year = None
            for look in rows[header_idx + 1 + offset + 1:]:
                if not look:
                    continue
                d = parse_date(get(look, 'date'))
                if d:
                    bb_year = int(d[:4])
                    break
            if bb_year is None:
                bb_year = current_fiscal_year
            result.append({
                'row_type':        'BEGINNING_BALANCE',
                'row_number':      excel_row_num,
                # The account is the enclosing section heading, never the
                # sentinel label -- "Beginning Balance" is not an account name,
                # and grouping the Trial Balance by it would strand every opening
                # balance under one bogus account.
                'account_name':    current_account_section or label_cell,
                'account_section': current_account_section,
                'description':     label_cell,
                'running_balance': last_numeric(row),
                'fiscal_year':     bb_year,
                # general_ledger_entries has no fiscal_year column (migration
                # 069) -- transaction_date IS the year key, and every downstream
                # read filters on it by date range. Leaving this None would keep
                # opening balances out of the Trial Balance even though they now
                # parse correctly. An opening balance is dated to the first
                # instant of the year it opens.
                'transaction_date': (f'{bb_year}-01-01' if bb_year else None),
                'raw_row_json':    raw_row_json,
            })
            continue

        # ── TOTAL ROW ─────────────────────────────────────────────────────────
        if first_lc.startswith('total'):
            result.append({
                'row_type':        'TOTAL_ROW',
                'row_number':      excel_row_num,
                'account_name':    first_cell,
                'account_section': current_account_section,
                'description':     first_cell,
                'running_balance': last_numeric(row),
                'fiscal_year':     current_fiscal_year,
                'transaction_date': None,
                'raw_row_json':    raw_row_json,
            })
            continue

        # ── TRANSACTION ───────────────────────────────────────────────────────
        if date_str:
            fiscal_year         = int(date_str[:4])
            current_fiscal_year = fiscal_year

            # account_name = the canonical posting account (new schema)
            if is_by_account:
                account_name = current_account_section or ''
                account_num  = ''
            else:
                account_name = cell_str(get(row, 'account'))
                account_num  = cell_str(get(row, 'account_num'))

            fiscal_month = int(date_str[5:7]) if date_str and len(date_str) >= 7 else None

            # Raw signed amount: prefer single Amount col; fall back to debit/credit
            raw_amount = parse_amount(get(row, 'amount'))
            if raw_amount is None:
                d = parse_amount(get(row, 'debit'))  or 0
                c = parse_amount(get(row, 'credit')) or 0
                if d != 0 or c != 0:
                    raw_amount = d - c

            if col['debit'] >= 0 or col['credit'] >= 0:
                debit_amount  = parse_amount(get(row, 'debit'))  or 0
                credit_amount = parse_amount(get(row, 'credit')) or 0
            elif raw_amount is not None:
                debit_amount  = raw_amount if raw_amount >= 0 else 0
                credit_amount = abs(raw_amount) if raw_amount < 0 else 0
            else:
                debit_amount = credit_amount = 0

            result.append({
                'row_type':          'TRANSACTION',
                'row_number':        excel_row_num,
                'transaction_date':  date_str,
                'fiscal_year':       fiscal_year,
                'fiscal_month':      fiscal_month,
                'account_section':   current_account_section,
                'account_name':      account_name,
                'account_number':    account_num or None,
                'transaction_number': cell_str(get(row, 'ref'))   or None,
                'transaction_type':  cell_str(get(row, 'type'))   or None,
                'memo':              cell_str(get(row, 'desc'))   or None,
                'split_account':     cell_str(get(row, 'split'))  or None,
                'amount':            raw_amount,
                'debit_amount':      debit_amount,
                'credit_amount':     credit_amount,
                'running_balance':   parse_amount(get(row, 'balance')),
                'raw_row_json':      raw_row_json,
            })
            continue

        # ── ACCOUNT HEADER ────────────────────────────────────────────────────
        # Any non-blank, non-date row that isn't a total or beginning-balance
        # is treated as an account section header.
        if first_cell:
            current_account_section = first_cell
            result.append({
                'row_type':        'ACCOUNT_HEADER',
                'row_number':      excel_row_num,
                'account_name':    first_cell,
                'account_section': first_cell,
                'description':     None,
                'running_balance': None,
                'fiscal_year':     current_fiscal_year,
                'transaction_date': None,
                'raw_row_json':    raw_row_json,
            })

    # detected_years only from TRANSACTION rows (the ones that have fiscal_year from a date)
    detected_years = sorted({
        r['fiscal_year'] for r in result
        if r.get('row_type') == 'TRANSACTION' and r.get('fiscal_year')
    })
    dbg(f'Extracted {len(result)} GL rows (all types), years={detected_years}')
    return {'rows': result, 'detected_years': detected_years}


# ── BANK STATEMENT ────────────────────────────────────────────────────────────

def extract_bank_statement(wb):
    rows = get_rows(wb, r'bank|statement|transaction|checking|savings')
    if len(rows) < 2:
        emit_error('No data rows found in bank statement file')

    BANK_DATE_ALIASES = DATE_ALIASES + ['transaction date', 'posting date', 'value date']
    BANK_DESC_ALIASES = DESC_ALIASES + ['description', 'transaction', 'payee', 'particulars']
    BANK_AMT_ALIASES  = AMOUNT_ALIASES + ['amount', 'transaction amount', 'withdrawal', 'deposit']
    BANK_BAL_ALIASES  = BALANCE_ALIASES + ['available balance', 'ledger balance']
    BANK_TYPE_ALIASES = TYPE_ALIASES + ['transaction type', 'code']

    def bank_scorer(row):
        score = 0
        for cell in row:
            h = lc(cell)
            if any(h == a or a in h for a in BANK_DATE_ALIASES[:4]):
                score += 4
            if any(h == a or a in h for a in BANK_DESC_ALIASES[:3]):
                score += 3
            if h in ('amount', 'debit', 'credit', 'withdrawal', 'deposit'):
                score += 3
        return score

    header_idx, _ = find_best_header_row(rows, bank_scorer, max_scan=20)
    header_row = rows[header_idx]

    col = {
        'date':    find_col(header_row, BANK_DATE_ALIASES),
        'desc':    find_col(header_row, BANK_DESC_ALIASES),
        'amount':  find_col(header_row, BANK_AMT_ALIASES),
        'debit':   find_col(header_row, DEBIT_ALIASES),
        'credit':  find_col(header_row, CREDIT_ALIASES),
        'balance': find_col(header_row, BANK_BAL_ALIASES),
        'type':    find_col(header_row, BANK_TYPE_ALIASES),
        'ref':     find_col(header_row, REF_ALIASES),
    }

    # Try to pull statement date and bank name from title rows above header
    statement_date = None
    bank_name      = None
    for i in range(min(header_idx, 10)):
        row_text = ' '.join(cell_str(c) for c in rows[i])
        if not statement_date:
            m = re.search(
                r'(january|february|march|april|may|june|july|august'
                r'|september|october|november|december)\s+\d{1,2},?\s+\d{4}',
                row_text, re.IGNORECASE,
            )
            if m:
                try:
                    statement_date = datetime.strptime(
                        m.group(0).replace(',', '').strip(), '%B %d %Y'
                    ).strftime('%Y-%m-%d')
                except ValueError:
                    pass

    # Date column probing fallback — same approach as GL extractor
    if col['date'] < 0:
        for c_idx in range(min(8, len(header_row))):
            for r_idx in range(header_idx + 1, min(header_idx + 6, len(rows))):
                if r_idx < len(rows) and c_idx < len(rows[r_idx]):
                    if parse_date(rows[r_idx][c_idx]):
                        col['date'] = c_idx
                        dbg(f'Date column probed at index {c_idx}')
                        break
            if col['date'] >= 0:
                break
    if col['date'] < 0:
        emit_error('Could not detect date column in bank statement file')

    # If no explicit amount/debit/credit column found, fall back to last numeric value per row
    use_last_numeric = col['amount'] < 0 and col['debit'] < 0 and col['credit'] < 0
    if use_last_numeric:
        dbg('No amount/debit/credit column detected — will use last numeric value per row')

    def get(row, key):
        c = col[key]
        return row[c] if c >= 0 and c < len(row) else None

    result     = []
    all_dates  = []

    for row in rows[header_idx + 1:]:
        if not row or all(v is None or str(v).strip() == '' for v in row):
            continue

        date_str = parse_date(get(row, 'date'))
        if not date_str:
            continue

        all_dates.append(date_str)

        # Determine amount and transaction type
        amount   = None
        txn_type = None

        raw_amount = get(row, 'amount')
        if raw_amount is not None:
            amount = parse_amount(raw_amount)

        if amount is None and col['debit'] >= 0:
            debit  = parse_amount(get(row, 'debit')  or 0)
            credit = parse_amount(get(row, 'credit') or 0)
            if debit and debit != 0:
                amount   = -abs(debit)
                txn_type = 'Withdrawal'
            elif credit and credit != 0:
                amount   = abs(credit)
                txn_type = 'Deposit'

        # Last-resort: use rightmost numeric cell after the date column
        if amount is None and use_last_numeric:
            date_col = col['date']
            for c in range(len(row) - 1, date_col, -1):
                v = parse_amount(row[c] if c < len(row) else None)
                if v is not None:
                    amount = v
                    break

        if amount is None:
            continue

        if not txn_type:
            txn_type = cell_str(get(row, 'type')) or None

        stmt_year  = int(date_str[:4])
        stmt_month = f'{date_str[:7]}-01'

        balance = parse_amount(get(row, 'balance'))

        result.append({
            'statement_date':  date_str,           # updated after loop if needed
            'statement_month': stmt_month,
            'bank_account':    bank_name or 'Bank Account',
            'bank_name':       bank_name,
            'account_type':    None,
            'transaction_date': date_str,
            'description':     cell_str(get(row, 'desc'))  or None,
            'reference':       cell_str(get(row, 'ref'))   or None,
            'amount':          amount,
            'transaction_type': txn_type,
            'running_balance': balance,
            'statement_year':  stmt_year,
        })

    # Finalize statement_date from max date if not found in title rows
    if all_dates and not statement_date:
        statement_date = max(all_dates)
        stmt_month     = f'{statement_date[:7]}-01'
        for r in result:
            r['statement_date']  = statement_date
            r['statement_month'] = stmt_month

    detected_years = sorted({r['statement_year'] for r in result})
    dbg(f'Extracted {len(result)} bank rows, years={detected_years}')
    return {'rows': result, 'detected_years': detected_years}


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Extract financial data from Excel files')
    parser.add_argument('--type', required=True,
        choices=['profit_loss', 'balance_sheet', 'general_ledger', 'bank_statement'])
    parser.add_argument('--filename', default='', help='Original filename (for logging)')
    args = parser.parse_args()

    dbg(f'type={args.type} filename="{args.filename}"')
    wb = load_wb()

    try:
        if args.type == 'profit_loss':
            result = extract_profit_loss(wb)
        elif args.type == 'balance_sheet':
            result = extract_balance_sheet(wb)
        elif args.type == 'general_ledger':
            result = extract_general_ledger(wb)
        else:
            result = extract_bank_statement(wb)
    except SystemExit:
        raise
    except Exception as e:
        import traceback
        traceback.print_exc(file=sys.stderr)
        emit_error(str(e))
        return

    wb.close()

    if not result['rows']:
        emit_error(f'No {args.type} rows extracted from file')

    emit(result)


if __name__ == '__main__':
    main()
